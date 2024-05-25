import polib
import pandas as pd
from tqdm import tqdm
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

# 读取和解析po文件
def read_po(file_path):
    po = polib.pofile(file_path)
    entries = {}
    for entry in po:
        key = (entry.msgid, entry.msgctxt)
        entries[key] = entry.msgstr
    return entries

# 合并两个po文件的内容
def merge_po_files(simplified_po, traditional_po):
    merged_entries = []
    
    all_keys = list(OrderedDict.fromkeys(list(simplified_po.keys()) + list(traditional_po.keys())))
    for key in tqdm(all_keys):
        msgid, msgctxt = key
        simp_translation = simplified_po.get(key, '')
        trad_translation = traditional_po.get(key, '')
        merged_entries.append([msgctxt, msgid, trad_translation, simp_translation])
    
    return merged_entries

# 导出为Excel表格
def export_to_excel(data, output_path):
    df = pd.DataFrame(data, columns=['msgctxt', 'msgid',  '繁中譯文', '简中译文'])
    df.to_excel(output_path, index=False)

    # 加载工作簿并获取活动工作表
    wb = load_workbook(output_path)
    ws = wb.active

    # 设置列宽
    column_widths = {'A': 32, 'B': 64, 'C': 64, 'D': 64}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 启用筛选
    ws.auto_filter.ref = ws.dimensions

    # 启用自动换行
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:

            if cell.column_letter == 'A':  # 'A' 是 msgctxt 列
                cell.alignment = Alignment(wrap_text=False)
            else:
                cell.alignment = Alignment(wrap_text=True)

    # 保存更改
    wb.save(output_path)

# 文件路径
simplified_po_path = './Translations/zh_HANS.po'
traditional_po_path = './Translations/zh_TW.po'
output_excel_path = 'merged_translations.xlsx'

# 读取po文件
simplified_entries = read_po(simplified_po_path)
traditional_entries = read_po(traditional_po_path)

# 合并内容
merged_entries = merge_po_files(simplified_entries, traditional_entries)

# 导出为Excel表格
export_to_excel(merged_entries, output_excel_path)

print(f"合并后的翻译已导出到 {output_excel_path}")

